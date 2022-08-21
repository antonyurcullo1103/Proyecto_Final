# -*- coding: utf-8 -*-
"""
Created on Mon Aug 14 21:25:52 2022

@author:
    ADAN MAIKON TERAN JUAREZ
    ANTONY CRISTIAN URCULLO ROSALES
    VICTOR ERNESTO DE LAS MONTAÑAS ORTEGA LIJERON
    RONALD TORRICO OVANDO
"""

# Importar dependencias
import openpyxl
import glob
import os
from tkinter import Tk, ttk, filedialog, messagebox, Button, Label, Menu, Entry, PhotoImage, NO, W, END, StringVar
from PIL import Image, ImageTk

from cargador_kaggle import CargadorKaggle
from dataset_processes import DatasetProcesses
from manejadorimagenes.imagen import Imagen

class Interfaz:

    NOMBRE_SISTEMA = "Sistema de Gestión de Imágenes"
    dataset_absolute_path = ""
    dataset_process = DatasetProcesses()
    imagen_actual: Imagen = None
    redimensionado = Image.open('NoImagen.png')
    
    def __init__(self):

        # Crear la ventana principal
        self.root = Tk()
        self.root.title(self.NOMBRE_SISTEMA)
        self.root.geometry("1360x720")

        # Crear el panel Treeview y añadirle a la ventana principal
        self.my_tree = ttk.Treeview(self.root)
        # Posicion treview(datagrid) en interfaz
        self.my_tree.pack()
        self.my_tree['columns'] = ("FILE NAME", "FORMAT", "SIZE", "URL")
        self.my_tree.column("#0", width=0, stretch=NO)
        self.my_tree.column("FILE NAME", anchor=W, width=200)
        self.my_tree.column("FORMAT", anchor=W, width=200)
        self.my_tree.column("SIZE", anchor=W, width=200)
        self.my_tree.column("URL", anchor=W, width=200)
        self.my_tree.grid(row=1, column=4)
        self.my_tree.bind('<Double-1>', self.get_fila)
        
        # Interfaz GUI(root) del Sistema
        # Crea Menubar
        self.menubar = Menu(self.root)

        # Agrega DatosKaggle Menu y command(llamada de funcion)
        self.datoskaggle = Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label='Obtener Datos', menu=self.datoskaggle)
        self.datoskaggle.add_command(label='Covid Kaggle', command=self.cargar_dataset_kaggle)

        # Agrega Nosotros a Menu and commands
        ayuda = Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label='Ayuda', menu=ayuda)
        ayuda.add_command(label='Sobre Nosotros', command=self.integrantes)

        # Titulo del Sistemas
        self.title_label = Label(self.root, text=self.NOMBRE_SISTEMA, font=('Arial bold', 25), bd=2)
        self.title_label.grid(row=0, column=2, columnspan=4, padx=5, pady=5)

        # label atributos
        self.file_name_label = Label(self.root, text="File Name", font=('Arial bold', 15))
        self.file_format_label = Label(self.root, text="Format", font=('Arial bold', 15))
        self.file_size_label = Label(self.root, text="Size", font=('Arial bold', 15))
        self.file_url_label = Label(self.root, text="Url", font=('Arial bold', 15))
        self.file_name_label.grid(row=2, column=0, padx=10, pady=10)
        self.file_format_label.grid(row=3, column=0, padx=10, pady=10)
        self.file_size_label.grid(row=4, column=0, padx=10, pady=10)
        self.file_url_label.grid(row=5, column=0, padx=10, pady=10)

        # imagen vacia
        self.img = PhotoImage(file='Noimagen.png')
        self.image_label = Button(self.root, image=self.img, borderwidth=2, relief='ridge',command=self.cambiar_Imagen)
        self.image_label.grid(row=1, column=1,columnspan=3)
        # Dando variables a entradas de texto
        self.t1 = StringVar()
        self.t2 = StringVar()
        self.t3 = StringVar()
        self.t4 = StringVar()
        # entradas atributos
        self.file_name_entry = Entry(self.root, textvariable=self.t1, width=25, bd=5, font=('Arial bold', 15))
        self.file_format_entry = Entry(self.root, textvariable=self.t2, width=25, bd=5, font=('Arial bold', 15))
        self.file_size_entry = Entry(self.root, textvariable=self.t3, width=25, bd=5, font=('Arial bold', 15))
        self.file_url_entry = Entry(self.root, textvariable=self.t4, width=25, bd=5, font=('Arial bold', 15))
        self.file_name_entry.grid(row=2, column=1, columnspan=3, padx=5, pady=5)
        self.file_format_entry.grid(row=3, column=1, columnspan=3, padx=5, pady=5)
        self.file_size_entry.grid(row=4, column=1, columnspan=3, padx=5, pady=5)
        self.file_url_entry.grid(row=5, column=1, columnspan=3, padx=5, pady=5)

        # Sección de botones del formulario
        self.boton_limpiar = Button(
            self.root, text="Limpiar", padx=5, pady=5, width=5,
            bd=3, font=('Arial', 15), bg="#0099ff", command=self.limpiar_entradas)
        self.boton_limpiar.grid(row=6, column=0, columnspan=1)
        
        self.boton_agregar = Button(
            self.root, text="Agregar", padx=5, pady=5, width=5,
            bd=3, font=('Arial', 15), bg="#0099ff", command=self.agregar_registro)
        self.boton_agregar.grid(row=6, column=1, columnspan=1)

        self.boton_editar = Button(
            self.root, text="Editar", padx=5, pady=5, width=5,
            bd=3, font=('Arial', 15), bg="#ffff00", command=self.editar_registro)
        self.boton_editar.grid(row=6, column=2, columnspan=1)

        self.boton_eliminar = Button(
            self.root, text="Eliminar", padx=5, pady=5, width=5,
            bd=3, font=('Arial', 15), bg="#e62e00", command=self.eliminar_registro)
        self.boton_eliminar.grid(row=6, column=3, columnspan=1)

        self.style = ttk.Style()
        self.style.configure("Treeview.Heading", font=('Arial bold', 15))

        # Combobox de categoria
        self.combo_box_categoria = ttk.Combobox(
            self.root, values=["Cargue Dataset desde Menú"],
            state="readonly",
            width=25,
            font=('Arial', 15))
        self.combo_box_categoria.grid(row=2, column=4, columnspan=3)
        # Trae excel a traves de la funcion selection_changed()
        self.combo_box_categoria.bind("<<ComboboxSelected>>", self.selection_changed)
        # Crear un status bar
        self.status_bar = Label(self.root, text="Estado actual ...", font=('Arial', 12))
        self.status_bar.grid(row=3, column=4, padx=0, pady=0)

        # Menuarriba
        self.root.config(menu=self.menubar)

    def get_fila(self, event):
        categoria = self.combo_box_categoria.get()
        item = self.my_tree.item(self.my_tree.focus())
        self.t1.set(item['values'][0])
        self.t2.set(item['values'][1])
        self.t3.set(item['values'][2])
        self.t4.set(item['values'][3])
        self.traer_imagen(self.t1.get(), categoria)
        
    def traer_imagen(self, file_name, cat):
        try:
            ruta_imagen = os.path.join(self.dataset_absolute_path, cat)
            ruta_imagen = os.path.join(ruta_imagen, "images")
            nombre_imagen = file_name + ".png"
            ruta_imagen = os.path.join(ruta_imagen, nombre_imagen)
            imagenTraida = Image.open(ruta_imagen)
            redimensionado = imagenTraida.resize((256,256))            
            render = ImageTk.PhotoImage(redimensionado)
            self.image_label.configure(image=render)
            self.image_label.image = render
        except ValueError:
            messagebox.showinfo("Alerta","Imagen no puede ser abierto... verifica de nuevo")
        except FileNotFoundError:
            messagebox.showinfo("Alerta","Imagen no encontrado...actualiza la imagen")

    def cargar_dataset_kaggle(self):
        self.status_bar.config(text = "Iniciando carga de dataset desde Kaggle ...")
        cargador = CargadorKaggle("dataset")
        self.dataset_absolute_path = cargador.download()
         
        # self.dataset_absolute_path = "C://Users//ronald torrico//Documents//ProyectoFinal//COVID-19_Radiography_Dataset"
        # Cargar categorias desde el folder de dataset
        categorias = self.dataset_process.loadCategoriesFromRootPath(self.dataset_absolute_path)
        values = list()
        for value in categorias:
            values.append(value.get_name())
        # Mostrar las categorias en la interfaz
        self.combo_box_categoria["values"] = values
        self.status_bar.config(text = "Lista de categorias actualizada")

    # Funcion abrir imagen
    def cambiar_Imagen(self):
        archivo = filedialog.askopenfilename(
            title="abrir", filetypes=[("Archivos png", "*.png")])
        archivo2 = Image.open(archivo)
        redimensionado = archivo2.resize((250, 250))
        
        render = ImageTk.PhotoImage(redimensionado)
        self.image_label.configure(image=render)
        self.image_label.image = render
        
        # Update entries
        self.imagen_actual = Imagen(archivo2.filename)
        self.imagen_actual.generate_metadata()
        # Actualizar los campos de texto
        self.file_name_entry.delete(0, END)
        self.file_format_entry.delete(0, END)
        self.file_size_entry.delete(0, END)
        self.file_name_entry.insert(0, self.imagen_actual.get_filename())
        self.file_format_entry.insert(0, self.imagen_actual.get_format())
        self.file_size_entry.insert(0, self.imagen_actual.get_size())
        # Mostrar metadatos
        messagebox.showinfo("Metadatos de la imagen", self.imagen_actual.__str__())
    
    def limpiar_entradas(self):
        self.image_label.configure(image=self.img)
        self.image_label.image = self.img
        
        self.imagen_actual = None
        self.file_name_entry.delete(0, END)
        self.file_format_entry.delete(0, END)
        self.file_size_entry.delete(0, END)
    
    # FUNCIONES
    # Funcion para mostrar en forma descendente las tuplas
    def reverse(self, tuples):
        new_tup = tuples[::-1]
        return new_tup
    
    # Funcion mostrar combobox seleccionado
    def selection_changed(self, event):
        selection = self.combo_box_categoria.get()
        if selection == "Cargue Dataset desde Menú":
            self.status_bar.config(text = "Debe cargar un dataset desde el menú 'Obtener datos'")
        else:
            self.file_openbox(selection)
    
    # Funcion file_open(excel) desde el combobox
    def file_openbox(self, seleccion):
        # Iterar los archivos excel en el folder de dataset
        dataset_files = glob.glob(os.path.join(self.dataset_absolute_path, "*.xlsx"))
        dataset_file = ""
        for dfile in dataset_files:
            if seleccion in dfile:
                dataset_file = dfile
                break
        
        self.cargar_tree_view(dataset_file)
            
    def cargar_tree_view(self, archivo_dataset):
        cat_dataframe = self.dataset_process.parseDataset(archivo_dataset)    
        # limpiar antiguo treeview(dataframe)
        self.clear_tree()
        # Establecer nuevo treeview(dataframe)
        self.my_tree['column'] = list(cat_dataframe.columns)
        self.my_tree['show'] = 'headings'
        # Iterar en las listas de columnas
        for column in self.my_tree['column']:
            self.my_tree.heading(column, text=column)

        # Colocar datos en treeview
        df_rows = cat_dataframe.to_numpy().tolist()
        for row in self.reverse(df_rows):
            self.my_tree.insert('', 'end', values=row)

    # Funcion limpiar filas del treeview
    def clear_tree(self):
        self.my_tree.delete(*self.my_tree.get_children())
    
    # Funcion agregar registro al excel desde el BOTON
    
    def agregar_registro(self):
        categoria = self.combo_box_categoria.get()
        archivo_excel = categoria + ".metadata.xlsx"
        wb = openpyxl.load_workbook(os.path.join(self.dataset_absolute_path, archivo_excel))
        ws = wb['Sheet1']

        file_name = str(self.file_name_entry.get())
        file_format = str(self.file_format_entry.get())
        file_size = str(self.file_size_entry.get())
        file_url = str(self.file_url_entry.get())
        was_found = False
        
        if file_name in ("", " "):
            print("Error Inserting Id")
        if file_format in ("", " "):
            print("Error Inserting Name")
        if file_size in ("", " "):
            print("Error Inserting Price")
        if file_url in ("", " "):
            print("Error Inserting Quantity")
        else:
           # Revisa si existe registro por FILE NAME
            for i in range(2, (ws.max_row)+1):
                if file_name == ws['A'+str(i)].value:
                    was_found = True
                    break                
    
            if was_found == True:
                messagebox.showinfo("Error", "File Name ya Existe!")
    
            else:
                lastRow = str((ws.max_row)+1)
                ws['A'+lastRow] = file_name
                ws['B'+lastRow] = file_format
                ws['C'+lastRow] = file_size
                ws['D'+lastRow] = file_url
                
                # Guardar la imagen en el dataset
                self.guardar_imagen(file_name, categoria)
                messagebox.showinfo("Alerta", "Datos Modificados Exitosamente!")
    
            wb.save(os.path.join(self.dataset_absolute_path, archivo_excel))

        # Volvemos a cargar los datos
        self.cargar_tree_view(os.path.join(self.dataset_absolute_path, archivo_excel))
            
    # Funcion editar registro del excel desde el BOTON
    def editar_registro(self):
        categoria= self.combo_box_categoria.get()
        archivo_excel = categoria + ".metadata.xlsx"
        wb = openpyxl.load_workbook(os.path.join(self.dataset_absolute_path, archivo_excel))
        ws = wb['Sheet1']
     
        file_name = str(self.file_name_entry.get())
        file_format = str(self.file_format_entry.get())
        file_size = str(self.file_size_entry.get())
        file_url = str(self.file_url_entry.get())
        was_found = False
        if file_name in ("", " "):
            messagebox.showinfo("Error", "Inserte File Name!")
        if file_format in ("", " "):
            messagebox.showinfo("Error", "Inserte Format!")
        if file_size in ("", " "):
            messagebox.showinfo("Error", "Inserte Size!")
        if file_url in ("", " "):
            messagebox.showinfo("Error", "Inserte Url!")
        else:
           #Revisa si existe registro por FILE NAME
            for i in range(2,(ws.max_row)+1):
                if file_name==ws['A'+str(i)].value:
                    was_found = True
                    ws['A'+str(i)]=file_name
                    ws['B'+str(i)]=file_format
                    ws['C'+str(i)]=file_size
                    ws['D'+str(i)]=file_url
                    
                    # Guardar imagen en el dataset
                    self.guardar_imagen(file_name, categoria)
                    # Guardar la imagen en el dataset
                    wb.save(os.path.join(self.dataset_absolute_path, archivo_excel))
                    
                    messagebox.showinfo("Alerta", "Datos Modificados Exitosamente!")
                    break

            if was_found == False:
                messagebox.showinfo("Error", "File Name no Existe!")

        # Volvemos a cargar los datos
        self.cargar_tree_view(os.path.join(self.dataset_absolute_path, archivo_excel))
            
    # Funcion eliminar registro del excel desde el BOTON
    def eliminar_registro(self):
        categoria= self.combo_box_categoria.get()
        archivo_excel = categoria + ".metadata.xlsx"
        wb = openpyxl.load_workbook(os.path.join(self.dataset_absolute_path, archivo_excel))
        ws = wb['Sheet1']
     
        file_name = str(self.file_name_entry.get())
        file_format = str(self.file_format_entry.get())
        file_size = str(self.file_size_entry.get())
        file_url = str(self.file_url_entry.get())
        was_found = False
        if file_name in ("", " "):
            messagebox.showinfo("Error", "Inserte File Name!")
        if file_format in ("", " "):
            messagebox.showinfo("Error", "Inserte Format!")
        if file_size in ("", " "):
            messagebox.showinfo("Error", "Inserte Size!")
        if file_url in ("", " "):
            messagebox.showinfo("Error", "Inserte Url!")
        else:
           #Revisa si existe registro por FILE NAME
            for i in range(2,(ws.max_row)+1):
                if file_name==ws['A'+str(i)].value:
                    was_found = True
                    if messagebox.askyesno('Alerta','Estas seguro que deseas eliminar el registo?'):
                        ws.delete_rows(i, 1)
                        self.eliminar_imagen(file_name, categoria)
                        messagebox.showinfo("Alerta", "Datos Eliminados Exitosamente!")
                        # Guardar la imagen en el dataset
                        wb.save(os.path.join(self.dataset_absolute_path, archivo_excel))
                        break

            if was_found == False:
                messagebox.showinfo("Error", "File Name no Existe!")

        # Volvemos a cargar los datos
        self.cargar_tree_view(os.path.join(self.dataset_absolute_path, archivo_excel))
            
    # Funcion mostrar Integrantes desde el Menubar
    def integrantes(self):
        messagebox.showinfo("Integrantes", "Adan Maikon Teran Juarez \n"
                            "Ronald Torrico \n"
                            "Victor Ernesto Ortega L \n"
                            "Antony Urcullo Rosales")

    #Guarda la imagen desde la intefaz con el nombre dado
    def guardar_imagen(self, file_name, cat):
        if self.redimensionado:
            ruta_imagen = os.path.join(self.dataset_absolute_path, cat)
            ruta_imagen = os.path.join(ruta_imagen, "images")
            nombre_imagen = file_name + ".png"
            ruta_imagen = os.path.join(ruta_imagen, nombre_imagen)
            self.redimensionado.save(ruta_imagen)            

    def eliminar_imagen(self, file_name,cat):
        if self.redimensionado:
            ruta_imagen = os.path.join(self.dataset_absolute_path, cat)
            ruta_imagen = os.path.join(ruta_imagen, "images")
            nombre_imagen = file_name + ".png"
            ruta_imagen = os.path.join(ruta_imagen, nombre_imagen)
            os.remove(ruta_imagen)
    
# Llama a la ventana principal
inter = Interfaz()
if __name__ == "__main__":
    inter.root.mainloop()
